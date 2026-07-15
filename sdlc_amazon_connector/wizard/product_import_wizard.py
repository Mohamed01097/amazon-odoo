import base64
import csv
import io
import logging

from odoo import models, fields
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


class AmazonProductImportWizard(models.TransientModel):
    _name = 'amazon.product.import.wizard'
    _description = 'Amazon Product Import Wizard'

    instance_id = fields.Many2one('amazon.instance', string='Amazon Instance', required=True)
    file = fields.Binary('File (CSV/Excel)', required=True)
    file_name = fields.Char('File Name')
    file_type = fields.Selection([
        ('csv', 'CSV'),
        ('excel', 'Excel'),
    ], string='File Type', default='csv', required=True)
    operation = fields.Selection([
        ('map', 'Map Products (Amazon SKU → Odoo Product)'),
        ('import', 'Import New Products'),
        ('update_price', 'Update Prices'),
        ('update_stock', 'Update Stock Levels'),
    ], string='Operation', default='map', required=True)
    delimiter = fields.Selection([
        (',', 'Comma (,)'),
        ('\t', 'Tab'),
        (';', 'Semicolon (;)'),
    ], string='Delimiter', default=',')

    def action_import(self):
        self.ensure_one()
        if not self.file:
            raise UserError("Please upload a file.")

        data = base64.b64decode(self.file)
        rows = self._parse_file(data)

        if self.operation == 'map':
            return self._process_mapping(rows)
        elif self.operation == 'import':
            return self._process_import(rows)
        elif self.operation == 'update_price':
            return self._process_price_update(rows)
        elif self.operation == 'update_stock':
            return self._process_stock_update(rows)

    def _parse_file(self, data):
        """Parse CSV or Excel file and return list of dicts."""
        if self.file_type == 'excel':
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(data))
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
                return rows
            except ImportError:
                raise UserError("openpyxl library is required for Excel import. Install with: pip install openpyxl")
        else:
            text = self._decode_uploaded_bytes(data)
            delimiter = self.delimiter or ','
            # See amazon_api.fetch_report_rows: splitlines() splits on lone \r
            # inside fields, silently corrupting rows. Normalize CRLF first
            # then strip stray CRs to a space.
            normalized = text.replace('\r\n', '\n').replace('\r', ' ')
            reader = csv.DictReader(io.StringIO(normalized), delimiter=delimiter)
            return list(reader)

    @staticmethod
    def _decode_uploaded_bytes(data):
        # Try UTF-8 (with BOM) first, then Excel's common Windows-1252, finally
        # Latin-1 with replace so we never raise UnicodeDecodeError for a user upload.
        for encoding in ('utf-8-sig', 'cp1252'):
            try:
                return data.decode(encoding)
            except UnicodeDecodeError:
                continue
        return data.decode('latin-1', errors='replace')

    def _process_mapping(self, rows):
        """Map Amazon products to Odoo products by SKU or barcode.
        Expected columns: sku (or seller-sku), odoo_product_id (or internal_reference or barcode)
        """
        mapped = 0
        errors = []
        for row in rows:
            sku = (row.get('sku') or row.get('seller-sku') or row.get('SKU') or '').strip()
            odoo_ref = (row.get('odoo_product_id') or row.get('internal_reference')
                        or row.get('default_code') or row.get('barcode') or '').strip()

            if not sku:
                continue

            amazon_prod = self.env['amazon.product'].search([
                ('sku', '=', sku),
                ('instance_id', '=', self.instance_id.id),
            ], limit=1)

            if not amazon_prod:
                errors.append("SKU not found: %s" % sku)
                continue

            if odoo_ref:
                odoo_prod = self.env['product.product'].search([
                    '|', ('default_code', '=', odoo_ref), ('barcode', '=', odoo_ref)
                ], limit=1)
                if odoo_prod:
                    amazon_prod.odoo_product_id = odoo_prod.id
                    mapped += 1
                else:
                    errors.append("Odoo product not found for ref: %s" % odoo_ref)

        msg = "%d product(s) mapped." % mapped
        if errors:
            msg += " Errors: %s" % "; ".join(errors[:10])
        return self._notify(msg, bool(errors))

    def _process_import(self, rows):
        """Import new Amazon products.
        Expected columns: sku, name, asin, price, quantity
        """
        imported = 0
        for row in rows:
            sku = (row.get('sku') or row.get('seller-sku') or row.get('SKU') or '').strip()
            if not sku:
                continue

            existing = self.env['amazon.product'].search([
                ('sku', '=', sku),
                ('instance_id', '=', self.instance_id.id),
            ], limit=1)

            raw_fc = (row.get('fulfillment_channel') or row.get('fulfillment-channel') or '').strip().upper()
            fc = 'AFN' if raw_fc in ('AFN', 'AMAZON_NA', 'AMAZON_EU', 'AMAZON_FE', 'AMAZON_IN') else 'MFN'
            vals = {
                'name': (row.get('name') or row.get('item-name') or row.get('product_name') or sku).strip(),
                'sku': sku,
                'asin': (row.get('asin') or row.get('asin1') or '').strip(),
                'amazon_price': float(row.get('price') or row.get('item-price') or 0),
                'amazon_qty': float(row.get('quantity') or row.get('qty') or 0),
                'brand': (row.get('brand') or '').strip(),
                'description': (row.get('description') or row.get('item-description') or '').strip(),
                'instance_id': self.instance_id.id,
                'last_sync_date': fields.Datetime.now(),
                'fulfillment_channel': fc,
            }
            if existing:
                existing.write(vals)
            else:
                self.env['amazon.product'].create(vals)
            imported += 1

        return self._notify("%d product(s) imported." % imported)

    def _process_price_update(self, rows):
        """Update prices on Amazon products.
        Expected columns: sku, price
        """
        updated = 0
        for row in rows:
            sku = (row.get('sku') or row.get('seller-sku') or row.get('SKU') or '').strip()
            price = row.get('price') or row.get('item-price') or row.get('new_price')
            if not sku or price is None:
                continue

            amazon_prod = self.env['amazon.product'].search([
                ('sku', '=', sku),
                ('instance_id', '=', self.instance_id.id),
            ], limit=1)
            if amazon_prod:
                amazon_prod.amazon_price = float(price)
                updated += 1

        return self._notify("%d price(s) updated." % updated)

    def _process_stock_update(self, rows):
        """Update stock quantities on Amazon products.
        Expected columns: sku, quantity
        """
        updated = 0
        for row in rows:
            sku = (row.get('sku') or row.get('seller-sku') or row.get('SKU') or '').strip()
            qty = row.get('quantity') or row.get('qty') or row.get('stock')
            if not sku or qty is None:
                continue

            amazon_prod = self.env['amazon.product'].search([
                ('sku', '=', sku),
                ('instance_id', '=', self.instance_id.id),
            ], limit=1)
            if amazon_prod:
                amazon_prod.amazon_qty = float(qty)
                updated += 1

        return self._notify("%d stock level(s) updated." % updated)

    def _notify(self, message, has_errors=False):
        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": "Product Import",
                "message": message,
                "type": "warning" if has_errors else "success",
                "sticky": has_errors,
            },
        }


class AmazonProductSetupWizard(models.TransientModel):
    _name = 'amazon.product.setup.wizard'
    _description = 'Amazon Product Initial Setup Wizard'

    instance_id = fields.Many2one('amazon.instance', string='Amazon Instance', required=True)
    operation = fields.Selection([
        ('link_existing', 'Link only existing Odoo products by SKU'),
        ('create_missing', 'Create missing Odoo products'),
    ], string='Operation', default='link_existing', required=True)
    update_odoo_prices_from_amazon = fields.Boolean(
        string='Update Odoo prices from Amazon once',
        help="Use only during initial setup or intentional reconciliation. Normal product sync does not update Odoo prices.",
    )

    def action_apply(self):
        self.ensure_one()
        domain = [('instance_id', '=', self.instance_id.id)]
        if not self.update_odoo_prices_from_amazon:
            domain.append(('odoo_product_id', '=', False))
        products = self.env['amazon.product'].search(domain)
        if not products:
            return self._notify("No Amazon products found for the selected operation.", True)

        if self.operation == 'link_existing':
            return products._setup_odoo_products(
                create_missing=False,
                update_prices_from_amazon=self.update_odoo_prices_from_amazon,
            )
        return products._setup_odoo_products(
            create_missing=True,
            update_prices_from_amazon=self.update_odoo_prices_from_amazon,
        )

    def _notify(self, message, has_errors=False):
        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": "Product Initial Setup",
                "message": message,
                "type": "warning" if has_errors else "success",
                "sticky": has_errors,
            },
        }
